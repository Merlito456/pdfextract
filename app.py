import streamlit as st
import pandas as pd
import pdfplumber
import re
from io import BytesIO
import base64
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(
    page_title="PO Details Data Extractor",
    page_icon="📄",
    layout="wide"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-title {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1F4E79;
        text-align: center;
        margin-bottom: 0.5rem;
    }
    .sub-title {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .preview-modal {
        position: fixed;
        top: 0;
        left: 0;
        width: 100%;
        height: 100%;
        background: rgba(0,0,0,0.5);
        z-index: 9999;
        display: flex;
        justify-content: center;
        align-items: center;
    }
    .preview-content {
        background: white;
        padding: 20px;
        border-radius: 10px;
        max-width: 90%;
        max-height: 90%;
        overflow: auto;
        box-shadow: 0 4px 20px rgba(0,0,0,0.3);
    }
    .preview-content table {
        width: 100%;
        border-collapse: collapse;
        font-family: 'Calibri', sans-serif;
    }
    .preview-content th {
        background-color: #1F4E79;
        color: white;
        padding: 10px;
        text-align: center;
        border: 1px solid #ddd;
        font-weight: bold;
    }
    .preview-content td {
        padding: 8px 12px;
        border: 1px solid #ddd;
        text-align: left;
    }
    .preview-content tr:nth-child(even) {
        background-color: #f9f9f9;
    }
    .preview-content tr:hover {
        background-color: #f5f5f5;
    }
    .close-btn {
        float: right;
        background: #ff4444;
        color: white;
        border: none;
        padding: 8px 16px;
        border-radius: 5px;
        cursor: pointer;
        font-size: 14px;
        margin-bottom: 10px;
    }
    .close-btn:hover {
        background: #cc0000;
    }
    .preview-btn {
        background-color: #4CAF50;
        color: white;
        padding: 10px 20px;
        border: none;
        border-radius: 5px;
        cursor: pointer;
        font-size: 14px;
        margin: 5px;
    }
    .preview-btn:hover {
        background-color: #45a049;
    }
    .stExpander {
        border: 1px solid #e0e0e0;
        border-radius: 5px;
    }
</style>
""", unsafe_allow_html=True)

# App title
st.markdown('<div class="main-title">📄 PO Details Data Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Extract Purchase Order line items from PDF to Excel</div>', unsafe_allow_html=True)

def extract_data_from_pdf(pdf_file):
    """
    Extract data from PDF file
    """
    extracted_data = []
    
    try:
        with pdfplumber.open(pdf_file) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text += text + "\n"
            
            # Parse the text
            return parse_pdf_text(all_text)
            
    except Exception as e:
        st.error(f"Error extracting data from {pdf_file.name}: {str(e)}")
        return []

def parse_pdf_text(text):
    """
    Parse the PDF text to extract line item data
    """
    parsed_rows = []
    
    # Clean up text
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    i = 0
    while i < len(lines):
        line = lines[i]
        
        # Look for line items - they start with "Line #No. Schedule Lines"
        if 'Line #No. Schedule Lines' in line or 'Line #No. Schedule LinesPart # / Description' in line.replace(' ', ''):
            i += 1
            continue
        
        # Look for data rows starting with line number
        match = re.search(r'^(\d+)\s+(?:Not\s+Available|Material)\s+([\d,]+(?:\.\d+)?\s*\([A-Z]+\))\s+\d+\s+[A-Za-z]+\s+\d+\s+([\d,]+\.\d+)\s+[A-Z]+\s+([\d,]+\.\d+)\s+[A-Z]+(?:\s+([\d,]+\.\d+)\s+[A-Z]+)?', line, re.IGNORECASE)
        
        if match:
            groups = match.groups()
            
            # Extract data
            line_num = groups[0].strip()
            qty = groups[1].strip() if len(groups) > 1 else ""
            unit_price = groups[2].strip() if len(groups) > 2 else ""
            subtotal = groups[3].strip() if len(groups) > 3 else ""
            
            # Look for product code and description
            pu = ""
            description = ""
            
            # Check next line for product code and description
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                
                # Try to extract PU and Description
                pu_match = re.search(r'^(\d+)[,\-]?\s*(.+)$', next_line)
                if pu_match:
                    pu = pu_match.group(1).strip()
                    description = pu_match.group(2).strip()
                else:
                    pu_match2 = re.search(r'^([A-Z0-9\-_]+)[,\s]+(.+)$', next_line)
                    if pu_match2:
                        pu = pu_match2.group(1).strip()
                        description = pu_match2.group(2).strip()
                    else:
                        pu_match3 = re.search(r'^([A-Z0-9\-_]+)', next_line)
                        if pu_match3:
                            potential_pu = pu_match3.group(1).strip()
                            if re.match(r'^\d+$', potential_pu) or re.match(r'^\d+[A-Z\-_]', potential_pu):
                                pu = potential_pu
                                description = next_line.replace(pu, '').strip()
                                if description.startswith(','):
                                    description = description[1:].strip()
                                if description.startswith('-'):
                                    description = description[1:].strip()
                            else:
                                description = next_line
                        else:
                            description = next_line
            
            # If we still don't have PU, try to extract from the line
            if not pu:
                pu_match_in_line = re.search(r'([A-Z0-9\-_]{6,})', line)
                if pu_match_in_line:
                    pu = pu_match_in_line.group(1)
            
            # Clean up description
            if description:
                description = re.sub(r'^[,:\-\s]+', '', description)
                description = re.sub(r'\s+', ' ', description).strip()
            
            # Add the row
            if line_num:
                parsed_rows.append({
                    'Line #': line_num,
                    'PU': pu,
                    'Description': description,
                    'QTY': qty,
                    'Unit Price': unit_price,
                    'Subtotal': subtotal
                })
                
                # Skip the description line if we used it
                if pu or description:
                    i += 1
        
        i += 1
    
    # If no data found with the main pattern, try alternative parsing
    if not parsed_rows:
        parsed_rows = parse_alternative(text)
    
    return parsed_rows

def parse_alternative(text):
    """
    Alternative parsing method for the data
    """
    parsed_rows = []
    
    # Split by line items
    lines = text.split('\n')
    
    current_row = {}
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Look for line number
        line_num_match = re.search(r'^(\d+)\s+(?:Not\s+Available|Material)', line)
        if line_num_match:
            # If we have a previous row, save it
            if current_row:
                parsed_rows.append(current_row)
            
            # Start new row
            current_row = {
                'Line #': line_num_match.group(1),
                'PU': '',
                'Description': '',
                'QTY': '',
                'Unit Price': '',
                'Subtotal': ''
            }
            
            # Extract QTY
            qty_match = re.search(r'([\d,]+(?:\.\d+)?\s*\([A-Z]+\))', line)
            if qty_match:
                current_row['QTY'] = qty_match.group(1).strip()
            
            # Extract prices
            price_matches = re.findall(r'([\d,]+\.\d+)\s+PHP', line)
            if len(price_matches) >= 2:
                current_row['Unit Price'] = price_matches[0]
                current_row['Subtotal'] = price_matches[1]
            elif len(price_matches) == 1:
                current_row['Subtotal'] = price_matches[0]
            
            # Check next line for PU and Description
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line and not next_line.startswith('STATUS') and 'Tax' not in next_line and 'Unconfirmed' not in next_line:
                    pu_match = re.search(r'^(\d+)[,\-]?\s*(.+)$', next_line)
                    if pu_match:
                        pu = pu_match.group(1).strip()
                        current_row['PU'] = pu
                        description = pu_match.group(2).strip()
                        if description.startswith(','):
                            description = description[1:].strip()
                        if description.startswith('-'):
                            description = description[1:].strip()
                        current_row['Description'] = description
                    else:
                        pu_match2 = re.search(r'^([A-Z0-9\-_]+)[,\s]+(.+)$', next_line)
                        if pu_match2:
                            pu = pu_match2.group(1).strip()
                            current_row['PU'] = pu
                            description = pu_match2.group(2).strip()
                            current_row['Description'] = description
                        else:
                            current_row['Description'] = next_line
                    i += 1
        
        i += 1
    
    # Add the last row
    if current_row:
        parsed_rows.append(current_row)
    
    return parsed_rows

def create_excel_with_style(df, filename):
    """
    Create Excel file with professional styling and formatting
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PO Details')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['PO Details']
        
        # Define styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Calibri', size=11)
        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply header styles
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Apply data styles
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = thin_border
                
                # Apply right alignment for number columns
                if col in [5, 6]:  # Unit Price and Subtotal columns
                    cell.alignment = number_alignment
                    # Format as number with 2 decimal places
                    if cell.value and isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                else:
                    cell.alignment = cell_alignment
        
        # Set column widths
        column_widths = {
            'A': 12,  # Line #
            'B': 15,  # PU
            'C': 50,  # Description
            'D': 15,  # QTY
            'E': 18,  # Unit Price
            'F': 18   # Subtotal
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
    
    return output.getvalue()

def create_combined_excel(all_dfs, filenames):
    """
    Create combined Excel file with multiple sheets
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for idx, (df, filename) in enumerate(zip(all_dfs, filenames)):
            # Get sheet name (truncate if too long)
            sheet_name = filename.rsplit('.', 1)[0][:31] if '.' in filename else filename[:31]
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + '...'
            
            # Make sheet name unique if duplicate
            if sheet_name in writer.sheets:
                sheet_name = f"{sheet_name[:28]}_{idx+1}"
            
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Get the worksheet
            worksheet = writer.sheets[sheet_name]
            
            # Apply styling
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cell_font = Font(name='Calibri', size=11)
            cell_alignment = Alignment(horizontal='left', vertical='center')
            number_alignment = Alignment(horizontal='right', vertical='center')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply header styles
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Apply data styles
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.border = thin_border
                    
                    if col in [5, 6]:
                        cell.alignment = number_alignment
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = cell_alignment
            
            # Set column widths
            column_widths = {'A': 12, 'B': 15, 'C': 50, 'D': 15, 'E': 18, 'F': 18}
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            worksheet.freeze_panes = 'A2'
    
    return output.getvalue()

def show_table_preview(df, title="Data Preview"):
    """
    Display a styled table preview in a modal-like format
    """
    # Convert DataFrame to HTML with styling
    styled_df = df.style.set_table_styles([
        {'selector': 'thead th', 'props': [('background-color', '#1F4E79'), ('color', 'white'), ('font-weight', 'bold'), ('padding', '10px'), ('text-align', 'center')]},
        {'selector': 'tbody td', 'props': [('padding', '8px 12px'), ('border', '1px solid #ddd')]},
        {'selector': 'tbody tr:nth-child(even)', 'props': [('background-color', '#f9f9f9')]},
        {'selector': 'tbody tr:hover', 'props': [('background-color', '#f5f5f5')]},
    ]).set_properties(**{
        'font-family': 'Calibri, sans-serif',
        'font-size': '11pt'
    }).format({
        'Unit Price': '{:,.2f}',
        'Subtotal': '{:,.2f}'
    }, na_rep='')
    
    html_table = styled_df.to_html()
    
    # Display in an expander as a modal-like view
    with st.expander(f"📊 {title}", expanded=True):
        st.markdown(f"""
        <div style="background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <div style="max-height: 600px; overflow: auto;">
                {html_table}
            </div>
            <div style="margin-top: 10px; color: #666; font-size: 12px;">
                Total items: {len(df)} | Columns: {', '.join(df.columns)}
            </div>
        </div>
        """, unsafe_allow_html=True)

def process_multiple_pdfs(uploaded_files, process_mode):
    """
    Process multiple PDFs based on selected mode
    """
    all_data = []
    file_names = []
    total_rows = 0
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, file in enumerate(uploaded_files):
        status_text.text(f"Processing: {file.name}...")
        data = extract_data_from_pdf(file)
        
        if data:
            df = pd.DataFrame(data)
            df = df.replace('', pd.NA).dropna(how='all')
            df = df.fillna('')
            df = df[~df['Line #'].astype(str).str.contains('Line', case=False, na=False)]
            
            desired_columns = ['Line #', 'PU', 'Description', 'QTY', 'Unit Price', 'Subtotal']
            existing_columns = [col for col in desired_columns if col in df.columns]
            if existing_columns:
                df = df[existing_columns]
            
            # Convert price columns to numeric for proper display
            for col in ['Unit Price', 'Subtotal']:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(',', '').str.replace(' PHP', '').str.replace('₱', '').str.replace(' ', '')
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            all_data.append(df)
            file_names.append(file.name)
            total_rows += len(df)
        
        progress_bar.progress((idx + 1) / len(uploaded_files))
    
    status_text.text(f"✅ Processed {len(uploaded_files)} files, extracted {total_rows} rows")
    progress_bar.empty()
    
    return all_data, file_names

# Main app UI
def main():
    st.subheader("📤 Upload PDF Files")
    
    # Allow multiple file uploads
    uploaded_files = st.file_uploader(
        "Choose PDF files (Purchase Orders)",
        type=['pdf'],
        accept_multiple_files=True,
        help="Upload one or multiple PO PDF files"
    )
    
    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)} file(s) uploaded successfully!")
        
        # Display uploaded files
        with st.expander("📋 Uploaded Files"):
            for file in uploaded_files:
                st.write(f"📄 {file.name} ({file.size / 1024:.1f} KB)")
        
        # Processing options
        st.subheader("⚙️ Processing Options")
        
        col1, col2 = st.columns(2)
        with col1:
            process_mode = st.radio(
                "Processing Mode",
                ["Combine into one", "Process individually"],
                help="Combine all PO data into one file or keep separate"
            )
        
        with col2:
            include_source = st.checkbox(
                "Include source filename in data",
                value=True,
                help="Add a column showing which PO file each row came from"
            )
        
        # Process button
        if st.button("🚀 Extract PO Data", type="primary"):
            all_data, file_names = process_multiple_pdfs(uploaded_files, process_mode)
            
            if all_data:
                st.subheader("📊 Extracted PO Details")
                
                if process_mode == "Combine into one":
                    # Combine all dataframes
                    combined_df = pd.concat(all_data, ignore_index=True)
                    
                    if include_source:
                        # Add source column
                        source_col = []
                        for df, filename in zip(all_data, file_names):
                            source_col.extend([filename] * len(df))
                        combined_df.insert(0, 'Source File', source_col)
                    
                    # Show preview button
                    col1, col2 = st.columns([1, 4])
                    with col1:
                        if st.button("👁️ Preview Table", key="preview_combined"):
                            show_table_preview(combined_df, "Combined PO Data Preview")
                    
                    # Display the dataframe
                    st.dataframe(combined_df, use_container_width=True)
                    
                    # Display statistics
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Line Items", len(combined_df))
                    with col2:
                        if 'Subtotal' in combined_df.columns and not combined_df['Subtotal'].empty:
                            subtotal_sum = combined_df['Subtotal'].sum()
                            if not pd.isna(subtotal_sum):
                                st.metric("Total PO Value", f"₱{subtotal_sum:,.2f}")
                            else:
                                st.metric("Total PO Value", "N/A")
                    with col3:
                        st.metric("POs Processed", len(uploaded_files))
                    
                    # Download combined file
                    st.subheader("📥 Download Extracted Data")
                    
                    # Create combined filename
                    base_name = "combined_po_details"
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        excel_data = create_excel_with_style(combined_df, base_name)
                        b64 = base64.b64encode(excel_data).decode()
                        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{base_name}.xlsx">📥 Download Combined Excel File</a>'
                        st.markdown(href, unsafe_allow_html=True)
                    
                    with col2:
                        csv = combined_df.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="📥 Download Combined CSV File",
                            data=csv,
                            file_name=f"{base_name}.csv",
                            mime="text/csv"
                        )
                    
                else:  # Process individually
                    # Show each file's data with preview buttons
                    for idx, (df, filename) in enumerate(zip(all_data, file_names)):
                        with st.expander(f"📄 {filename} ({len(df)} items)", expanded=False):
                            # Preview button for individual file
                            col1, col2 = st.columns([1, 4])
                            with col1:
                                if st.button(f"👁️ Preview", key=f"preview_{idx}"):
                                    show_table_preview(df, f"Preview: {filename}")
                            
                            st.dataframe(df, use_container_width=True)
                            
                            # Show mini stats for this file
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Line Items", len(df))
                            with col2:
                                if 'Subtotal' in df.columns:
                                    subtotal_sum = df['Subtotal'].sum()
                                    if not pd.isna(subtotal_sum):
                                        st.metric("PO Value", f"₱{subtotal_sum:,.2f}")
                                    else:
                                        st.metric("PO Value", "N/A")
                            with col3:
                                st.metric("Fields", len(df.columns))
                    
                    # Download options
                    st.subheader("📥 Download Extracted Data")
                    
                    download_option = st.radio(
                        "Download options:",
                        ["Download individual files", "Download all as one combined file"]
                    )
                    
                    if download_option == "Download individual files":
                        # Create download buttons for each file
                        for idx, (df, filename) in enumerate(zip(all_data, file_names)):
                            base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename
                            
                            col1, col2 = st.columns(2)
                            with col1:
                                excel_data = create_excel_with_style(df, base_name)
                                b64 = base64.b64encode(excel_data).decode()
                                href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{base_name}_PO_Details.xlsx">📥 Download {base_name}.xlsx</a>'
                                st.markdown(href, unsafe_allow_html=True)
                            
                            with col2:
                                csv = df.to_csv(index=False).encode('utf-8')
                                st.download_button(
                                    label=f"📥 Download {base_name}.csv",
                                    data=csv,
                                    file_name=f"{base_name}_PO_Details.csv",
                                    mime="text/csv",
                                    key=f"csv_{idx}"
                                )
                    else:
                        # Combine all into one file with multiple sheets
                        combined_excel = create_combined_excel(all_data, file_names)
                        b64 = base64.b64encode(combined_excel).decode()
                        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="combined_po_details.xlsx">📥 Download Combined Excel File (Multiple Sheets)</a>'
                        st.markdown(href, unsafe_allow_html=True)
                        
                        # Also provide combined CSV
                        combined_df = pd.concat(all_data, ignore_index=True)
                        if include_source:
                            source_col = []
                            for df, filename in zip(all_data, file_names):
                                source_col.extend([filename] * len(df))
                            combined_df.insert(0, 'Source File', source_col)
                        
                        # Preview for combined data
                        col1, col2 = st.columns([1, 4])
                        with col1:
                            if st.button("👁️ Preview Combined Data", key="preview_combined_individual"):
                                show_table_preview(combined_df, "Combined PO Data Preview")
                        
                        csv = combined_df.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="📥 Download Combined CSV File",
                            data=csv,
                            file_name="combined_po_details.csv",
                            mime="text/csv"
                        )
                
                # Show raw extracted data
                with st.expander("🔍 View Raw Extracted Data"):
                    st.json(all_data)
                
            else:
                st.warning("No data could be extracted from any of the PDF files.")
    else:
        st.info("👈 Please upload one or more PO PDF files to extract data")
        
        # Show instructions
        with st.expander("ℹ️ How to use this app"):
            st.markdown("""
            ### PO Details Data Extractor - PDF to Excel
            
            This app extracts line item details from Purchase Order (PO) PDF files and converts them to Excel format.
            
            **What it extracts:**
            - Line Number
            - Product Unit (PU)
            - Description
            - Quantity (QTY)
            - Unit Price
            - Subtotal
            
            **Features:**
            - 📤 Upload multiple PDF files at once
            - 🔄 Combine all POs into one file or keep separate
            - 👁️ Preview data before downloading
            - 📥 Download as Excel or CSV
            - 🎨 Professionally formatted Excel output
            
            **How to use:**
            1. Upload one or more PO PDF files
            2. Choose processing mode (Combine or Individual)
            3. Click "Extract PO Data"
            4. Preview the extracted data
            5. Download as Excel or CSV
            
            **Supported PDF Format:**
            The app is designed for PO PDFs that contain line items in a specific format with:
            - Line numbers
            - Product codes (PU)
            - Descriptions
            - Quantities in format like "3 (EA)"
            - Unit prices and subtotals
            """)

if __name__ == "__main__":
    main()